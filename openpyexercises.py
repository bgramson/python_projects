import openpyxl


wb=openpyxl.load_workbook("lastFile.xlsx")

sheets = wb.sheetnames
print("\n")
print("Number of sheets in the workbook are ",len(sheets))
print("\n")
print("Active sheet is: ",wb.active.title)

sh1 = wb['Populations - all use cases']

# print(type(sh1))
print("\n")
data1 =  sh1['A7'].value

data2 = wb['Populations - all use cases']['A9'].value

print(data1)
print("\n")
print(data2)

row = sh1.max_row
column = sh1.max_column
print("No of rows: ",row)
print("No of columns: ",column)

for i in range(1,row+1):
    for j in range(1,column+1):
        if i != None or j != None:
            print(sh1.cell(i,j=='').value)